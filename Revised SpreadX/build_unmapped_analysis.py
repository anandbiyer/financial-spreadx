"""Build Financials_Provided/Unmapped_Analysis.xlsx — a unified, single-worksheet
record of every unmapped BS/P&L line across the latest spread run per document.

For each unmapped item it records: the referenced annual report, the extraction
statement it came from, the model's TOP SUGGESTED CoA (these items are unmapped,
so there is NO accepted mapping — only the best candidate) + its score, and a
derived classification of Main line item / Sub-total / Total.

Re-runnable:  .venv\\Scripts\\python.exe build_unmapped_analysis.py
"""

from __future__ import annotations

import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from db.models import CoaReference, Document, UnmappedItem
from db.session import session_scope

_ROOT = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(_ROOT, "Financials_Provided", "Unmapped_Analysis.xlsx")

STMT_LABEL = {
    "balance_sheet": "Balance Sheet",
    "income_statement": "Income Statement",
    "equity_statement": "Statement of Changes in Equity",
}

# Classification of the extracted line by its label shape.
_TOTAL_RE = re.compile(r"\btotal\b|grand total|total assets less current", re.I)
_SUBTOTAL_RE = re.compile(
    r"\bnet\s+(assets|current|profit|loss|income|sales|revenue|worth|block)\b|"
    r"\bgross\s+(profit|block|fixed)\b|operating\s+(profit|loss|income)|"
    r"profit\s+(before|after|for|available)|loss\s+for\s+the|"
    r"shareholders.{0,3}\s*funds|sub-?total|comprehensive income|"
    r"\bEBIT\b|\bEBITDA\b|\bPBT\b|\bPAT\b|profit/\(loss\)",
    re.I,
)


def classify(label: str) -> str:
    if _TOTAL_RE.search(label):
        return "Total"
    if _SUBTOTAL_RE.search(label):
        return "Sub-total"
    return "Main line item"


def conf_band(s):
    if s is None:
        return "No suggestion"
    if s >= 0.55:
        return "Near-miss (0.55-0.59)"
    if s >= 0.45:
        return "Low (0.45-0.54)"
    if s >= 0.30:
        return "Very low (0.30-0.44)"
    return "No fit (<0.30)"


# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1A1917")
HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=9)
BOLD = Font(name="Calibri", size=9, bold=True)
GREY = PatternFill("solid", fgColor="D9D9D9")
TYPE_FILL = {"Total": PatternFill("solid", fgColor="BDD7EE"),
             "Sub-total": PatternFill("solid", fgColor="DDEBF7")}
BAND_FILL = {
    "Near-miss (0.55-0.59)": PatternFill("solid", fgColor="C6EFCE"),
    "Low (0.45-0.54)": PatternFill("solid", fgColor="FFEB9C"),
    "Very low (0.30-0.44)": PatternFill("solid", fgColor="FCD5B4"),
    "No fit (<0.30)": PatternFill("solid", fgColor="F2DCDB"),
}


def main() -> None:
    with session_scope() as s:
        coa_name = {c.coa_id: c.line_item_name
                    for c in s.execute(select(CoaReference)).scalars()}
        # latest document per filename
        latest: dict[str, tuple[str, object]] = {}
        for d in s.execute(select(Document)).scalars():
            if not d.filename.lower().endswith(".pdf"):
                continue
            if d.filename not in latest or d.created_at > latest[d.filename][1]:
                latest[d.filename] = (d.id, d.created_at)

        records = []
        for fn in sorted(latest):
            doc_id = latest[fn][0]
            items = s.execute(
                select(UnmappedItem).where(
                    UnmappedItem.document_id == doc_id,
                    UnmappedItem.status == "pending",
                )
            ).scalars().all()
            for it in items:
                sg = it.claude_suggestions or []
                best = max(sg, key=lambda c: (c.get("score") or 0)) if sg else {}
                scid = best.get("coa_id", "")
                score = best.get("score")
                vs = it.value_spread or {}
                yrs = sorted(vs.keys())
                records.append({
                    "report": os.path.splitext(fn)[0],
                    "statement": STMT_LABEL.get(it.statement_type, it.statement_type),
                    "line_item": it.raw_label or "",
                    "type": classify(it.raw_label or ""),
                    "coa_id": scid,
                    "coa_name": coa_name.get(scid, ""),
                    "score": score,
                    "band": conf_band(score),
                    "value": vs.get(yrs[-1]) if yrs else None,
                    "values": ", ".join(f"{y}={vs.get(y)}" for y in yrs),
                    "reason": it.ambiguity_note or "",
                    "status": it.status,
                })

    # sort: report → statement → highest suggestion score first
    records.sort(key=lambda r: (r["report"], r["statement"], -(r["score"] or 0)))

    headers = [
        "Annual Report (source)", "Extraction Statement", "Line Item (raw label)",
        "Line Item Type", "Top Suggested CoA ID (NOT an accepted mapping)",
        "Top Suggested CoA Name", "Suggestion Confidence", "Confidence Band",
        "Value (latest year)", "Values by year", "Reason / Ambiguity", "Status",
    ]
    widths = [34, 30, 42, 15, 20, 28, 12, 20, 16, 22, 78, 10]

    wb = Workbook()
    ws = wb.active
    ws.title = "Unmapped Analysis"
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w

    for r, rec in enumerate(records, 2):
        vals = [rec["report"], rec["statement"], rec["line_item"], rec["type"],
                rec["coa_id"], rec["coa_name"], rec["score"], rec["band"],
                rec["value"], rec["values"], rec["reason"], rec["status"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BOLD if (c == 4 and rec["type"] != "Main line item") else BODY
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (3, 11)))
        if rec["type"] in TYPE_FILL:
            ws.cell(row=r, column=4).fill = TYPE_FILL[rec["type"]]
        if rec["band"] in BAND_FILL:
            ws.cell(row=r, column=8).fill = BAND_FILL[rec["band"]]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"

    wb.save(OUT)
    print(f"Wrote {len(records)} unmapped records to {OUT}")
    # quick console recap
    from collections import Counter
    by_type = Counter(r["type"] for r in records)
    by_stmt = Counter(r["statement"] for r in records)
    print("By type:", dict(by_type))
    print("By statement:", dict(by_stmt))


if __name__ == "__main__":
    main()
