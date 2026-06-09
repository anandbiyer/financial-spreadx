"""Seed the CoA reference table from BS_PL_Line_Items_V1.xlsx (idempotent).

Run once:  python -m db.seed_coa

The spreadsheet has no sign_convention / is_subtotal / is_memo_item columns, so
those are DERIVED here from the line-item name + category (the spec's Section 03
heuristics): contra-assets and accumulated depreciation/amortisation/impairment
→ 'contra'; P&L reductions (returns, discounts, interest expense, tax, COGS,
operating expenses) → 'negative'; everything else → 'positive'.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from config import logger
from db.queries import count_coa_reference, upsert_coa_reference

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = _PROJECT_ROOT / "BS_PL_Line_Items_V1.xlsx"

# Expected header → model field.
_COLUMN_MAP = {
    "id": "coa_id",
    "line item name": "line_item_name",
    "statement": "statement",
    "broad category": "broad_category",
    "sub-category": "sub_category",
    "definition": "definition",
    "spreading guidance": "spreading_guidance",
}

# ── Derivation heuristics ─────────────────────────────────────────────────────

_CONTRA_RE = re.compile(
    r"\(-\)|accum(?:ulated)?\s+deprec|accum(?:ulated)?\s+amort|"
    r"bad\s+debt\s+reserve|accumulated\s+impairment|amort.*impairment|deprec.*impairment",
    re.IGNORECASE,
)

# P&L items that reduce profit/equity → negative sign convention.
_PL_NEGATIVE_RE = re.compile(
    r"\breturns?\b|\bdiscount|allowance|\bcogs\b|cost of goods|cost of sales|"
    r"\bexpense|expenditure|deprec|amort|impairment|interest\s+expense|"
    r"\btax\b|withdrawal|dividend|\bloss\b|provision\s+for|written?\s*off|write[- ]?off",
    re.IGNORECASE,
)
# Never treat these P&L items as negative even if a keyword above matches.
_PL_POSITIVE_GUARD_RE = re.compile(
    r"revenue|\bsales\b|gross\s+profit|net\s+income|interest\s+income|"
    r"\bgain\b|income\s+from|\bprofit\s+before\b|\bprofit\s+after\b|tax\s+credit",
    re.IGNORECASE,
)

_SUBTOTAL_RE = re.compile(
    r"\btotal\b|subtotal|gross\s+profit|net\s+revenue|net\s+sales|net\s+income|"
    r"net\s+profit|net\s+loss|profit\s+before|profit\s+after|operating\s+profit|"
    r"\bebit\b|\bebitda\b|\bpbt\b|\bpat\b",
    re.IGNORECASE,
)


def derive_sign_convention(name: str, statement: str) -> str:
    if _CONTRA_RE.search(name):
        return "contra"
    if statement.strip().upper() in ("P&L", "PL", "P & L"):
        if _PL_NEGATIVE_RE.search(name) and not _PL_POSITIVE_GUARD_RE.search(name):
            return "negative"
    return "positive"


def derive_is_subtotal(name: str) -> bool:
    return bool(_SUBTOTAL_RE.search(name))


def derive_is_memo(name: str) -> bool:
    return "memo" in name.lower()


# ── Parsing ───────────────────────────────────────────────────────────────────

def _parse_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    idx = {_COLUMN_MAP[h]: i for i, h in enumerate(header) if h in _COLUMN_MAP}
    if "coa_id" not in idx:
        return []

    entries: list[dict] = []
    for r in rows[1:]:
        coa_id = r[idx["coa_id"]]
        if coa_id is None or str(coa_id).strip() == "":
            continue
        e = {field: (r[i] if i < len(r) and r[i] is not None else "") for field, i in idx.items()}
        e["coa_id"] = str(e["coa_id"]).strip()
        e["line_item_name"] = str(e.get("line_item_name", "")).strip()
        e["statement"] = str(e.get("statement", "")).strip()
        name = e["line_item_name"]
        e["sign_convention"] = derive_sign_convention(name, e["statement"])
        e["is_subtotal"] = derive_is_subtotal(name)
        e["is_memo_item"] = derive_is_memo(name)
        entries.append(e)
    return entries


def load_coa_entries(xlsx_path: str | Path = DEFAULT_XLSX) -> list[dict]:
    """Parse all CoA entries from the workbook (no DB writes)."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    entries: list[dict] = []
    for ws in wb.worksheets:
        entries.extend(_parse_sheet(ws))
    wb.close()
    return entries


def seed_coa_reference(xlsx_path: str | Path = DEFAULT_XLSX) -> int:
    """Load + upsert all CoA entries. Returns the number seeded."""
    entries = load_coa_entries(xlsx_path)
    upsert_coa_reference(entries)
    return len(entries)


def main() -> None:
    if not DEFAULT_XLSX.exists():
        raise SystemExit(f"CoA reference file not found: {DEFAULT_XLSX}")
    n = seed_coa_reference()
    total = count_coa_reference()
    logger.info(f"[seed] upserted {n} CoA entries; coa_reference now has {total} rows")
    print(f"Seeded {n} CoA entries. coa_reference total = {total}")


if __name__ == "__main__":
    main()
