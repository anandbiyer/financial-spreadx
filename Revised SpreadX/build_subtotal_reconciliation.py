"""Build Financials_Provided/Subtotal_Reconciliation.xlsx — a single-worksheet
view of the subtotal cross-foot stored on each Document (reconciliation_result).

Each subtotal appears as a bold SUBTOTAL row (extracted vs computed, PASS/FAIL,
grouping confidence) followed by its component rows (mapping status, CoA, value,
sign-flip). Mirrors build_unmapped_analysis.py.

Re-runnable:  .venv\\Scripts\\python.exe build_subtotal_reconciliation.py
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from db.models import Document
from db.session import session_scope

_ROOT = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(_ROOT, "Financials_Provided", "Subtotal_Reconciliation.xlsx")

STMT_LABEL = {
    "balance_sheet": "Balance Sheet",
    "income_statement": "Income Statement",
    "equity_statement": "Statement of Changes in Equity",
}

HEADER_FILL = PatternFill("solid", fgColor="1A1917")
HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=9)
BOLD = Font(name="Calibri", size=9, bold=True)
GREEN = PatternFill("solid", fgColor="C6EFCE")   # pass
RED = PatternFill("solid", fgColor="F2DCDB")     # fail / unmapped
AMBER = PatternFill("solid", fgColor="FFEB9C")   # incomplete / sign flip


def _val(d, year):
    return (d or {}).get(year) if year else None


def main() -> None:
    with session_scope() as s:
        latest: dict[str, tuple] = {}
        for d in s.execute(select(Document)).scalars():
            if not d.filename.lower().endswith(".pdf"):
                continue
            if d.filename not in latest or d.created_at > latest[d.filename][1]:
                latest[d.filename] = (d.reconciliation_result, d.created_at)

        rows = []  # flat render list of (kind, dict)
        for fn in sorted(latest):
            recon = latest[fn][0]
            if not recon:
                continue
            report = fn[:-4] if fn.lower().endswith(".pdf") else fn
            for sub in recon.get("subtotals", []):
                rows.append(("subtotal", report, sub))
                for comp in sub.get("components", []):
                    rows.append(("component", report, (sub, comp)))

    headers = [
        "Annual Report", "Statement", "Kind", "Line Item",
        "Type / Status", "CoA ID", "CoA Name", "Confidence",
        "Value (latest yr)", "Computed Sum", "Difference", "Result",
        "Grouping Method", "Flags",
    ]
    widths = [30, 26, 11, 40, 16, 9, 24, 11, 16, 14, 14, 12, 11, 22]

    wb = Workbook()
    ws = wb.active
    ws.title = "Subtotal Reconciliation"
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w

    r = 2
    for kind, report, payload in rows:
        if kind == "subtotal":
            sub = payload
            yr = sub.get("headline_year")
            result = {True: "PASS", False: "FAIL", None: "INCOMPLETE"}[sub.get("pass")]
            flags = []
            if sub.get("has_unmapped_component"):
                flags.append("unmapped component")
            vals = [report, STMT_LABEL.get(sub["statement_type"], sub["statement_type"]),
                    "▣ SUBTOTAL", sub.get("raw_label", ""),
                    "Subtotal/Total", "", "", "",
                    _val(sub.get("extracted"), yr), _val(sub.get("computed"), yr),
                    _val(sub.get("difference"), yr), result,
                    sub.get("grouping_method", sub.get("grouping_confidence", "")),
                    "; ".join(flags)]
            fill = {"PASS": GREEN, "FAIL": RED, "INCOMPLETE": AMBER}[result]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = BOLD
                cell.alignment = Alignment(vertical="top", wrap_text=(c == 4))
            ws.cell(row=r, column=12).fill = fill
        else:
            sub, comp = payload
            yr = sub.get("headline_year")
            cflags = []
            if comp.get("sign_flipped"):
                cflags.append("SIGN FLIP")
            vals = [report, STMT_LABEL.get(sub["statement_type"], sub["statement_type"]),
                    "    └ component", comp.get("raw_label", ""),
                    comp.get("status", ""), comp.get("coa_id", ""), comp.get("coa_name", ""),
                    comp.get("confidence"), _val(comp.get("raw_values"), yr),
                    "", "", "", "", "; ".join(cflags)]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = BODY
                cell.alignment = Alignment(vertical="top", wrap_text=(c == 4))
            if comp.get("status") != "mapped":
                ws.cell(row=r, column=5).fill = RED
            if comp.get("sign_flipped"):
                ws.cell(row=r, column=14).fill = AMBER
        r += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(r - 1, 1)}"
    wb.save(OUT)

    n_sub = sum(1 for k, *_ in rows if k == "subtotal")
    print(f"Wrote {n_sub} subtotals ({r - 2} total rows) to {OUT}")


if __name__ == "__main__":
    main()
