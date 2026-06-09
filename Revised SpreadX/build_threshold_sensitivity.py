"""Build Financials_Provided/Threshold_Sensitivity.xlsx — an OFFLINE what-if study
of the COA-mapping confidence threshold (production default 0.60, applied at
spreading/coa_mapper.py: `result.confidence >= confidence_threshold`).

The threshold is a pure POST-HOC gate: it does not change the LLM call or its
output, only whether the model's top pick is accepted. Every unmapped item already
persists its top suggestion (claude_suggestions -> coa_id + score) and its raw
value_spread, so "what if the threshold were T" == "accept every unmapped item whose
top suggestion score >= T". This is fully reconstructable from spreadx.db with ZERO
LLM calls and no behavior change — this script writes NOTHING to the DB.

For each latest-per-filename document it reconstructs, at several thresholds:
  - coverage (mapped vs still-unmapped),
  - the Balance-Sheet identity A = L + E (reusing check_balance_sheet_identity),
  - subtotal reconciliation "unmapped-component" flags (recomputed from the stored
    reconciliation_result by joining flagged components to suggestion scores).

CAVEATS (also noted on the workbook's Notes sheet):
  * Confidence proxy — the gate uses result.confidence, but unmapped items persist the
    top candidate SCORE (not confidence). We use the top suggestion score as the proxy,
    the same convention build_unmapped_analysis.py uses. R12 "UNMAPPED" rows have no
    suggestion and are never recovered (correct).
  * Lower-only — all thresholds are <= 0.60, so baseline mapped rows are never removed;
    lowering only ADDS rows. At T=0.60 the reconstruction reproduces the stored
    balance_check_result (a built-in sanity check).
  * No correctness measure — lowering the threshold accepts the model's top pick, which
    may be wrong. The report quantifies coverage / balance movement / materiality, NOT
    accuracy. A balance difference that shrinks as T drops suggests the recovered
    mappings are plausibly right; one that grows suggests noise. See the "Flipped @ 0.20"
    sheet to eyeball quality.

Re-runnable:  .venv\\Scripts\\python.exe build_threshold_sensitivity.py
"""

from __future__ import annotations

import os
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from db.models import Document
from db.queries import get_all_coa_reference, get_coa_mappings_by_document
from db.session import session_scope
from db.models import UnmappedItem
from spreading.balance_checks import check_balance_sheet_identity
from spreading.sign_conventions import apply_sign_to_spread

_ROOT = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(_ROOT, "Financials_Provided", "Threshold_Sensitivity.xlsx")

# Editable: the gate values to simulate. Production default is 0.60.
THRESHOLDS = [0.60, 0.55, 0.45, 0.20, 0.0]

STMT_LABEL = {
    "balance_sheet": "Balance Sheet",
    "income_statement": "Income Statement",
    "equity_statement": "Statement of Changes in Equity",
}

# Confidence bands for the suggestion-score breakdown (mirrors build_unmapped_analysis).
_BANDS = [
    ("Near-miss (0.55-0.59)", 0.55),
    ("Low (0.45-0.54)", 0.45),
    ("Very low (0.30-0.44)", 0.30),
    ("Marginal (0.20-0.29)", 0.20),
    ("No fit (0.00-0.19)", 0.0),
]


def _band(score) -> str:
    if score is None:
        return "No suggestion"
    for label, lo in _BANDS:
        if score >= lo:
            return label
    return "No fit (0.00-0.19)"


def _best_suggestion(item, valid_ids=None) -> tuple[str, float | None]:
    """Top (coa_id, score) suggestion that points at a REAL CoA, or ("", None).

    Suggestions whose coa_id is the R12 "UNMAPPED" sentinel, empty, or a hallucinated
    id (when `valid_ids` is given) are ignored — the model declining is not a mapping
    that any threshold should recover.
    """
    sg = item.claude_suggestions or []
    usable = [c for c in sg
              if c.get("coa_id") and c.get("coa_id") != "UNMAPPED"
              and (valid_ids is None or c.get("coa_id") in valid_ids)]
    if not usable:
        return "", None
    best = max(usable, key=lambda c: (c.get("score") or 0))
    return best.get("coa_id", "") or "", best.get("score")


def _latest_year(spread: dict) -> str | None:
    yrs = sorted((spread or {}).keys())
    return yrs[-1] if yrs else None


def _norm(label: str) -> str:
    return str(label or "").strip().lower()


# ── per-document reconstruction ────────────────────────────────────────────────

def _load_docs(s):
    """Latest Document per filename (.pdf only), as ORM objects."""
    latest: dict[str, tuple[object, object]] = {}
    for d in s.execute(select(Document)).scalars():
        if not d.filename.lower().endswith(".pdf"):
            continue
        if d.filename not in latest or d.created_at > latest[d.filename][1]:
            latest[d.filename] = (d, d.created_at)
    return [latest[fn][0] for fn in sorted(latest)]


def _simulate(doc, coa_by_id, s) -> dict:
    """Reconstruct coverage / balance / reconciliation across THRESHOLDS for one doc."""
    valid_ids = set(coa_by_id)
    mappings = get_coa_mappings_by_document(doc.id)
    pending = s.execute(
        select(UnmappedItem).where(
            UnmappedItem.document_id == doc.id,
            UnmappedItem.status == "pending",
        )
    ).scalars().all()

    # Equity is not spread post-§3.9 (it is recorded as terminal not_spread, not a
    # threshold-recoverable row). These docs were last spread BEFORE that change, so
    # equity still sits here as "pending" — exclude it so the threshold lever is
    # studied on BS/P&L rows only, matching current behaviour.
    n_equity_excluded = sum(1 for u in pending if u.statement_type == "equity_statement")
    unmapped = [u for u in pending if u.statement_type != "equity_statement"]

    # Pre-extract each unmapped item's top (real-CoA) suggestion once.
    u_meta = []
    for u in unmapped:
        coa_id, score = _best_suggestion(u, valid_ids)
        u_meta.append({
            "raw_label": u.raw_label or "",
            "statement_type": u.statement_type or "",
            "value_spread": u.value_spread or {},
            "best_coa_id": coa_id,
            "best_score": score,
            "key": (_norm(u.raw_label), u.statement_type or ""),
        })

    recon = doc.reconciliation_result or {}
    recon_subtotals = recon.get("subtotals", []) if isinstance(recon, dict) else []
    baseline_with_unmapped = (recon.get("summary", {}) or {}).get("with_unmapped_component")

    baseline_mapped = len(mappings)
    total_unmapped = len(unmapped)

    per_threshold = []
    for t in THRESHOLDS:
        accepted = [m for m in u_meta
                    if m["best_coa_id"] and m["best_score"] is not None and m["best_score"] >= t]
        accepted_keys = {m["key"] for m in accepted}

        # Augmented mapped set for the balance identity (sign applied to recovered rows).
        augmented = [{"coa_id": m["coa_id"], "value_spread": m.get("value_spread") or {},
                      "raw_label": m.get("raw_label", "")} for m in mappings]
        for m in accepted:
            sign = (coa_by_id.get(m["best_coa_id"]) or {}).get("sign_convention", "positive")
            vs, _ = apply_sign_to_spread(m["value_spread"], sign)
            augmented.append({"coa_id": m["best_coa_id"], "value_spread": vs,
                              "raw_label": m["raw_label"]})

        balance = check_balance_sheet_identity(augmented, coa_by_id)

        # Recompute subtotal "has unmapped component" flags at this threshold: a component
        # is resolved if it was already mapped OR it joins to a now-accepted unmapped item.
        with_unmapped = None
        if recon_subtotals:
            cnt = 0
            for rep in recon_subtotals:
                stmt = rep.get("statement_type", "")
                still = False
                for comp in rep.get("components", []):
                    if comp.get("status") == "mapped":
                        continue
                    if (_norm(comp.get("raw_label")), stmt) in accepted_keys:
                        continue
                    still = True
                    break
                if still:
                    cnt += 1
            with_unmapped = cnt

        mapped_rows = baseline_mapped + len(accepted)
        still_unmapped = total_unmapped - len(accepted)
        denom = mapped_rows + still_unmapped
        value_recovered = 0.0
        for m in accepted:
            y = _latest_year(m["value_spread"])
            v = (m["value_spread"] or {}).get(y) if y else None
            if isinstance(v, (int, float)):
                value_recovered += abs(v)

        per_threshold.append({
            "threshold": t,
            "mapped": mapped_rows,
            "newly_accepted": len(accepted),
            "still_unmapped": still_unmapped,
            "coverage": (mapped_rows / denom) if denom else None,
            "is_balanced": balance.get("isBalanced"),
            "difference": balance.get("difference"),
            "total_assets": balance.get("totalAssets"),
            "total_le": balance.get("totalLiabilitiesAndEquity"),
            "primary_year": balance.get("primary_year"),
            "recon_with_unmapped": with_unmapped,
            "value_recovered": value_recovered,
        })

    bands = Counter(_band(m["best_score"]) for m in u_meta)

    # Items that newly map at T=0.20 (above the 0.60 baseline) — for quality eyeballing.
    flipped_020 = [m for m in u_meta
                   if m["best_coa_id"] and m["best_score"] is not None
                   and 0.20 <= m["best_score"] < 0.60]

    return {
        "doc": doc,
        "baseline_mapped": baseline_mapped,
        "total_unmapped": total_unmapped,
        "n_equity_excluded": n_equity_excluded,
        "baseline_with_unmapped": baseline_with_unmapped,
        "per_threshold": per_threshold,
        "bands": bands,
        "flipped_020": flipped_020,
    }


# ── styling ────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1A1917")
HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=9)
BOLD = Font(name="Calibri", size=9, bold=True)
DOC_FILL = PatternFill("solid", fgColor="D9D9D9")
BASE_FILL = PatternFill("solid", fgColor="DDEBF7")   # T = 0.60 baseline row
GOOD = PatternFill("solid", fgColor="C6EFCE")
BAD = PatternFill("solid", fgColor="F2DCDB")


def _hdr(ws, headers, widths, row=1):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w


def _pct(x):
    return f"{x * 100:.1f}%" if isinstance(x, (int, float)) else ""


def _num(x):
    return round(x, 2) if isinstance(x, (int, float)) else None


def main() -> None:
    with session_scope() as s:
        coa_by_id = {c["coa_id"]: c for c in get_all_coa_reference()}
        docs = _load_docs(s)
        results = [_simulate(d, coa_by_id, s) for d in docs]

        wb = Workbook()

        # ── Sheet 1: Threshold Comparison ─────────────────────────────────────
        ws = wb.active
        ws.title = "Threshold Comparison"
        headers = ["Document", "Threshold", "Mapped", "Newly accepted", "Still unmapped",
                   "Coverage", "Balanced?", "Balance diff (primary yr)", "Total Assets",
                   "Total L+E", "Subtotals w/ unmapped comp.", "Value recovered (latest yr)"]
        widths = [34, 10, 9, 14, 14, 10, 10, 22, 16, 16, 24, 22]
        _hdr(ws, headers, widths)
        r = 2
        for res in results:
            name = os.path.splitext(res["doc"].filename)[0]
            for i, pt in enumerate(res["per_threshold"]):
                vals = [
                    name if i == 0 else "",
                    pt["threshold"],
                    pt["mapped"], pt["newly_accepted"], pt["still_unmapped"],
                    _pct(pt["coverage"]),
                    "" if pt["is_balanced"] is None else ("YES" if pt["is_balanced"] else "no"),
                    _num(pt["difference"]), _num(pt["total_assets"]), _num(pt["total_le"]),
                    pt["recon_with_unmapped"], _num(pt["value_recovered"]),
                ]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.font = BOLD if c == 1 else BODY
                    cell.alignment = Alignment(vertical="top", wrap_text=(c == 1))
                if pt["threshold"] == 0.60:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=r, column=c).fill = BASE_FILL
                bcell = ws.cell(row=r, column=7)
                if pt["is_balanced"] is True:
                    bcell.fill = GOOD
                elif pt["is_balanced"] is False:
                    bcell.fill = BAD
                r += 1
            r += 1  # blank spacer between documents
        ws.freeze_panes = "A2"

        # ── Sheet 2: Band Breakdown ───────────────────────────────────────────
        ws2 = wb.create_sheet("Band Breakdown")
        band_order = [b[0] for b in _BANDS] + ["No suggestion"]
        headers2 = ["Document", "BS/P&L unmapped", "Equity excluded"] + band_order
        widths2 = [34, 16, 16] + [20] * len(band_order)
        _hdr(ws2, headers2, widths2)
        r = 2
        for res in results:
            name = os.path.splitext(res["doc"].filename)[0]
            vals = [name, res["total_unmapped"], res["n_equity_excluded"]] + \
                [res["bands"].get(b, 0) for b in band_order]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.font = BOLD if c == 1 else BODY
                cell.alignment = Alignment(vertical="top", wrap_text=(c == 1))
            r += 1
        # totals row
        tot = Counter()
        for res in results:
            tot.update(res["bands"])
        trow = ["TOTAL", sum(res["total_unmapped"] for res in results),
                sum(res["n_equity_excluded"] for res in results)] + \
               [tot.get(b, 0) for b in band_order]
        for c, v in enumerate(trow, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.font = BOLD
            cell.fill = DOC_FILL
        ws2.freeze_panes = "A2"

        # ── Sheet 3: Flipped @ 0.20 ───────────────────────────────────────────
        ws3 = wb.create_sheet("Flipped @ 0.20")
        headers3 = ["Document", "Statement", "Line Item (raw label)",
                    "Top Suggested CoA ID", "Top Suggested CoA Name", "Score",
                    "Value (latest yr)"]
        widths3 = [34, 28, 44, 18, 28, 8, 16]
        _hdr(ws3, headers3, widths3)
        r = 2
        for res in results:
            name = os.path.splitext(res["doc"].filename)[0]
            rows = sorted(res["flipped_020"], key=lambda m: -(m["best_score"] or 0))
            for m in rows:
                y = _latest_year(m["value_spread"])
                v = (m["value_spread"] or {}).get(y) if y else None
                vals = [name, STMT_LABEL.get(m["statement_type"], m["statement_type"]),
                        m["raw_label"], m["best_coa_id"],
                        (coa_by_id.get(m["best_coa_id"]) or {}).get("line_item_name", ""),
                        m["best_score"], _num(v) if isinstance(v, (int, float)) else None]
                for c, val in enumerate(vals, 1):
                    cell = ws3.cell(row=r, column=c, value=val)
                    cell.font = BODY
                    cell.alignment = Alignment(vertical="top", wrap_text=(c in (3, 5)))
                r += 1
        ws3.freeze_panes = "A2"

        # ── Sheet 4: Notes ────────────────────────────────────────────────────
        ws4 = wb.create_sheet("Notes")
        notes = [
            "Threshold Sensitivity — OFFLINE what-if study (no LLM calls, no DB writes).",
            "",
            "The confidence threshold is a post-hoc gate (spreading/coa_mapper.py): it only",
            "decides whether the model's top pick is accepted, not what the model returns.",
            "Lowering it to T == 'accept every unmapped item whose top suggestion score >= T',",
            "reconstructed here from the persisted suggestions + raw values.",
            "",
            "CAVEATS:",
            " * Confidence proxy — unmapped items persist the top candidate SCORE, not the",
            "   gate's result.confidence. We use the score as the proxy (same convention as",
            "   build_unmapped_analysis.py). R12 'UNMAPPED' rows have no suggestion and are",
            "   never recovered.",
            " * Lower-only — all thresholds <= 0.60, so baseline mapped rows are never removed.",
            "   The T=0.60 row reproduces the stored balance_check_result (a sanity check).",
            " * No correctness measure — lower thresholds accept the model's top pick, which",
            "   may be wrong. A balance diff that SHRINKS as T drops suggests the recovered",
            "   mappings are plausibly right; one that GROWS suggests noise. Use the",
            "   'Flipped @ 0.20' sheet to eyeball mapping quality.",
            " * Reconciliation PASS/FAIL is threshold-independent (the foot sums raw values);",
            "   only the 'subtotals with an unmapped component' count moves.",
            " * Equity excluded — equity_statement is not spread (it is recorded as terminal",
            "   not_spread). These docs were last spread BEFORE that change, so equity still",
            "   sits as 'pending'; it is excluded here so the threshold lever is studied on",
            "   BS/P&L rows only. Baseline 'mapped' still reflects the last persisted run.",
            " * How to read the balance column — A=L+E is OFF for every doc at every threshold",
            "   (a known extraction/mapping completeness issue, not a code defect). The signal",
            "   is the DIRECTION of the difference as T drops: shrinking => recovered mappings",
            "   are plausibly right; growing => noise.",
            f"Thresholds simulated: {THRESHOLDS}",
        ]
        for i, line in enumerate(notes, 1):
            cell = ws4.cell(row=i, column=1, value=line)
            cell.font = BOLD if (line and not line.startswith(" ") and line.endswith(")")
                                 or line.startswith("Threshold Sensitivity")) else BODY
        ws4.column_dimensions["A"].width = 100

    wb.save(OUT)

    # ── console recap ──────────────────────────────────────────────────────────
    print(f"Wrote threshold sensitivity for {len(results)} document(s) to {OUT}\n")
    for res in results:
        name = os.path.splitext(res["doc"].filename)[0]
        print(f"{name}: baseline {res['baseline_mapped']} mapped / "
              f"{res['total_unmapped']} BS&P&L unmapped "
              f"({res['n_equity_excluded']} equity excluded)")
        for pt in res["per_threshold"]:
            bal = "—" if pt["is_balanced"] is None else ("balanced" if pt["is_balanced"] else "OFF")
            diff = f"{pt['difference']:,.0f}" if isinstance(pt["difference"], (int, float)) else "—"
            print(f"   T={pt['threshold']:.2f}  cover={_pct(pt['coverage']):>6}  "
                  f"+{pt['newly_accepted']:>3} accepted  A=L+E {bal:>8} (diff {diff})  "
                  f"recon-unmapped-subtotals={pt['recon_with_unmapped']}")
        print()


if __name__ == "__main__":
    main()
