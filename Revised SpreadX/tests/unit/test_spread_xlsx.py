"""Unit tests for the Spread XLSX export — focused on the Unmapped Items sheet,
which guarantees every extracted BS/P&L line is visible in the workbook (mapped
or unmapped) for manual review.
"""

from __future__ import annotations

from io import BytesIO

import openpyxl

from export.spread_xlsx import build_spread_xlsx

_FORMATTED = {
    "years": ["2023", "2024"],
    "balance_sheet": [
        {"coa_id": "BS-001", "line_item_name": "Cash", "raw_label": "Cash + Bank",
         "value_spread": {"2024": 100.0}, "confidence": 0.97, "mapping_source": "claude",
         "source_extraction_ids": [2, 5]},
    ],
    "pl": [],
}
_MAPPINGS = [
    {"coa_id": "BS-001", "raw_label": "Cash + Bank", "statement_type": "balance_sheet",
     "confidence": 0.97, "mapping_source": "claude", "aggregated_from": 2, "rationale": "x",
     "source_extraction_ids": [2, 5]},
]
_UNMAPPED = [
    {"raw_label": "Amounts falling due within one year", "statement_type": "balance_sheet",
     "value_spread": {"2023": -1317667, "2024": -1374630},
     "claude_suggestions": [{"coa_id": "BS-086", "score": 0.58, "reason": "aggregate"},
                            {"coa_id": "BS-070", "score": 0.45, "reason": "trade"}],
     "ambiguity_note": "Highly aggregated UK GAAP creditors line", "status": "pending"},
]


def _load(data: bytes):
    return openpyxl.load_workbook(BytesIO(data))


_RECON = {
    "subtotals": [
        {"statement_type": "balance_sheet", "raw_label": "Net current assets",
         "headline_year": "2024", "extracted": {"2024": 70}, "computed": {"2024": 70},
         "difference": {"2024": 0}, "pass": True, "grouping_method": "absorbed",
         "has_unmapped_component": True,
         "components": [
             {"raw_label": "Current assets", "is_subtotal": True, "status": "unmapped",
              "coa_id": "", "coa_name": "", "confidence": None, "raw_values": {"2024": 100},
              "sign_flipped": False},
             {"raw_label": "Creditors", "is_subtotal": False, "status": "mapped",
              "coa_id": "BS-079", "coa_name": "Other Accruals", "confidence": 0.7,
              "raw_values": {"2024": -30}, "sign_flipped": False},
         ]},
    ],
    "summary": {"total": 1, "passed": 1, "failed": 0, "incomplete": 0},
}


def test_workbook_has_all_seven_sheets():
    wb = _load(build_spread_xlsx(_FORMATTED, _MAPPINGS, _UNMAPPED, reconciliation=_RECON))
    assert wb.sheetnames == ["Balance Sheet", "P&L", "Unmapped Items",
                             "Subtotal Reconciliation", "Confidence & Source",
                             "Learned Mappings Applied", "Run Usage & Cost"]


def test_extraction_ids_traced_onto_sheets():
    wb = _load(build_spread_xlsx(_FORMATTED, _MAPPINGS, _UNMAPPED, reconciliation=_RECON))
    bs = wb["Balance Sheet"]
    assert bs.cell(row=1, column=4).value == "Extraction ID(s)"
    assert bs.cell(row=2, column=4).value == "2, 5"  # aggregated source rows
    conf = wb["Confidence & Source"]
    assert conf.cell(row=1, column=3).value == "Extraction ID(s)"
    assert conf.cell(row=2, column=3).value == "2, 5"
    unm = wb["Unmapped Items"]
    assert unm.cell(row=1, column=3).value == "Extraction ID(s)"


def test_usage_sheet_renders_when_present():
    usage = {
        "by_stage": {"extraction": {"calls": 3, "input_tokens": 1000, "output_tokens": 200,
                                    "cache_read": 0, "cache_creation": 0, "cost_usd": 0.006},
                     "spreading": {"calls": 2, "input_tokens": 500, "output_tokens": 100,
                                   "cache_read": 0, "cache_creation": 0, "cost_usd": 0.003}},
        "total": {"calls": 5, "input_tokens": 1500, "output_tokens": 300,
                  "cache_read": 0, "cache_creation": 0, "cost_usd": 0.009},
        "models": ["claude-sonnet-4-6"], "unknown_model_pricing": False, "pricing_note": "x",
    }
    wb = _load(build_spread_xlsx(_FORMATTED, _MAPPINGS, _UNMAPPED, usage=usage))
    ws = wb["Run Usage & Cost"]
    flat = [c for row in ws.iter_rows(values_only=True) for c in row]
    assert "extraction" in flat and "spreading" in flat and "TOTAL" in flat
    assert 1500 in flat  # total input tokens


def test_reconciliation_sheet_renders_subtotal_and_components():
    wb = _load(build_spread_xlsx(_FORMATTED, _MAPPINGS, _UNMAPPED, reconciliation=_RECON))
    ws = wb["Subtotal Reconciliation"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    sub = rows[0]
    assert "Net current assets" in sub and "PASS" in sub and "absorbed" in sub
    # flag counts only the non-subtotal unmapped component (Creditors mapped, the
    # absorbed subtotal excluded) -> 0 missing leaves here.
    assert "missing leaf" not in (sub[-1] or "")
    comp_labels = [r[2] for r in rows[1:]]
    assert "Current assets" in comp_labels and "Creditors" in comp_labels


def test_reconciliation_sheet_empty_state():
    wb = _load(build_spread_xlsx(_FORMATTED, _MAPPINGS, _UNMAPPED))
    ws = wb["Subtotal Reconciliation"]
    assert "No subtotals" in str(ws.cell(row=2, column=1).value)


def test_unmapped_line_and_suggestions_present():
    wb = _load(build_spread_xlsx(_FORMATTED, _MAPPINGS, _UNMAPPED))
    ws = wb["Unmapped Items"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    row = rows[0]
    assert row[0] == "Amounts falling due within one year"
    assert row[1] == "balance_sheet"
    # value columns for 2023/2024 carry the extracted figures
    assert -1374630 in row
    # top suggestion + score, then alternatives, reason, status
    assert "BS-086" in row
    assert 0.58 in row
    assert any("BS-070" in str(c) for c in row)
    assert row[-1] == "pending"


def test_unmapped_sheet_empty_message_when_all_mapped():
    wb = _load(build_spread_xlsx(_FORMATTED, _MAPPINGS, []))
    ws = wb["Unmapped Items"]
    assert "No unmapped items" in str(ws.cell(row=2, column=1).value)


def test_unmapped_items_default_none_is_safe():
    # Backward-compatible call (no unmapped_items) must still build the sheet.
    wb = _load(build_spread_xlsx(_FORMATTED, _MAPPINGS))
    assert "Unmapped Items" in wb.sheetnames
