"""Spread XLSX export — 7 sheets (SPR-002).

Sheet 1: Balance Sheet (CoA order)
Sheet 2: P&L (CoA order)
Sheet 3: Unmapped Items (extracted BS/P&L lines that did NOT get a confident CoA)
Sheet 4: Subtotal Reconciliation (each subtotal cross-footed vs its components)
Sheet 5: Confidence & Source (every mapped row)
Sheet 6: Learned Mappings Applied (attribution audit trail)
Sheet 7: Run Usage & Cost (LLM token usage + estimated cost, per stage)

Sheets 3 & 4 make the workbook self-contained for manual review: sheet 3 shows
every extracted BS/P&L line as mapped (sheets 1/2) or unmapped, and sheet 4 shows
whether each extracted subtotal foots against the sum of its component lines
(surfacing mis-mapped, mis-signed, or missing components). Cash-flow lines have no
CoA target and remain only in the raw extraction workbook.

Confidence cells are colour-coded green/amber/orange per the Section 04 bands.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _ext_ids(row: dict) -> str:
    """Comma-joined sorted source extraction ids for traceability."""
    ids = row.get("source_extraction_ids") or []
    return ", ".join(str(i) for i in sorted(ids))

HEADER_FILL = PatternFill(start_color="1A1917", end_color="1A1917", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=9)
SUBTOTAL_FONT = Font(name="Calibri", size=9, bold=True)

_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_AMBER = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_ORANGE = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")


def _confidence_fill(conf: float | None):
    if conf is None:
        return None
    if conf >= 0.90:
        return _GREEN
    if conf >= 0.75:
        return _AMBER
    if conf >= 0.60:
        return _ORANGE
    return None


def _write_header(ws, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _write_statement_sheet(ws, rows: list[dict], years: list[str]) -> None:
    headers = ["CoA ID", "Line Item", "Raw Label (source)", "Extraction ID(s)"] \
        + [f"Value {y}" for y in years] + ["Confidence", "Source"]
    _write_header(ws, headers)
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 14

    r = 2
    for row in rows:
        font = SUBTOTAL_FONT if row.get("is_subtotal") else BODY_FONT
        ws.cell(row=r, column=1, value=row["coa_id"]).font = font
        ws.cell(row=r, column=2, value=row["line_item_name"]).font = font
        ws.cell(row=r, column=3, value=row.get("raw_label", "")).font = BODY_FONT
        ws.cell(row=r, column=4, value=_ext_ids(row)).font = BODY_FONT
        col = 5
        for y in years:
            val = (row.get("value_spread") or {}).get(y)
            ws.cell(row=r, column=col, value=val).font = BODY_FONT
            col += 1
        conf = row.get("confidence")
        conf_cell = ws.cell(row=r, column=col, value=conf)
        conf_cell.font = BODY_FONT
        fill = _confidence_fill(conf)
        if fill:
            conf_cell.fill = fill
        col += 1
        ws.cell(row=r, column=col, value=(row.get("mapping_source") or "")).font = BODY_FONT
        r += 1


_UNMAPPED_FILL = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")


def _write_unmapped_sheet(ws, unmapped_items: list[dict], years: list[str]) -> None:
    """List every extracted BS/P&L line that did not receive a confident CoA
    mapping (routed to the unmapped queue), with its best suggestion and reason —
    so the workbook shows what was missed, not just what mapped."""
    headers = ["Raw Label (source)", "Statement Type", "Extraction ID(s)"] \
        + [f"Value {y}" for y in years] \
        + ["Top Suggested CoA", "Top Score", "Other Suggestions", "Reason / Ambiguity", "Status"]
    _write_header(ws, headers)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14  # Extraction ID(s)
    top_coa_idx = 4 + len(years)  # first column after the year columns
    ws.column_dimensions[get_column_letter(top_coa_idx)].width = 18  # Top Suggested CoA
    ws.column_dimensions[get_column_letter(top_coa_idx + 3)].width = 80  # Reason / Ambiguity

    if not unmapped_items:
        ws.cell(row=2, column=1,
                value="(No unmapped items — every extracted BS/P&L line was mapped.)").font = BODY_FONT
        return

    for r, item in enumerate(unmapped_items, 2):
        raw_cell = ws.cell(row=r, column=1, value=item.get("raw_label", ""))
        raw_cell.font = BODY_FONT
        raw_cell.fill = _UNMAPPED_FILL
        ws.cell(row=r, column=2, value=item.get("statement_type", "")).font = BODY_FONT
        ws.cell(row=r, column=3, value=_ext_ids(item)).font = BODY_FONT
        col = 4
        vs = item.get("value_spread") or {}
        for y in years:
            ws.cell(row=r, column=col, value=vs.get(y)).font = BODY_FONT
            col += 1
        sugg = item.get("claude_suggestions") or []
        top = sugg[0] if sugg else {}
        ws.cell(row=r, column=col, value=top.get("coa_id", "")).font = BODY_FONT
        col += 1
        ws.cell(row=r, column=col, value=top.get("score")).font = BODY_FONT
        col += 1
        others = "; ".join(f"{s.get('coa_id')} ({s.get('score')})" for s in sugg[1:])
        ws.cell(row=r, column=col, value=others).font = BODY_FONT
        col += 1
        ws.cell(row=r, column=col, value=item.get("ambiguity_note", "")).font = BODY_FONT
        col += 1
        ws.cell(row=r, column=col, value=item.get("status", "pending")).font = BODY_FONT


def _write_confidence_sheet(ws, coa_mappings: list[dict]) -> None:
    headers = ["CoA ID", "Raw Label", "Extraction ID(s)", "Statement Type",
               "Confidence", "Source", "Aggregated From", "Rationale"]
    _write_header(ws, headers)
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["H"].width = 80
    for r, m in enumerate(coa_mappings, 2):
        ws.cell(row=r, column=1, value=m.get("coa_id", "")).font = BODY_FONT
        ws.cell(row=r, column=2, value=m.get("raw_label", "")).font = BODY_FONT
        ws.cell(row=r, column=3, value=_ext_ids(m)).font = BODY_FONT
        ws.cell(row=r, column=4, value=m.get("statement_type", "")).font = BODY_FONT
        conf_cell = ws.cell(row=r, column=5, value=m.get("confidence"))
        conf_cell.font = BODY_FONT
        fill = _confidence_fill(m.get("confidence"))
        if fill:
            conf_cell.fill = fill
        ws.cell(row=r, column=6, value=m.get("mapping_source", "")).font = BODY_FONT
        ws.cell(row=r, column=7, value=m.get("aggregated_from", 1)).font = BODY_FONT
        ws.cell(row=r, column=8, value=m.get("rationale", "")).font = BODY_FONT


def _write_learned_sheet(ws, coa_mappings: list[dict]) -> None:
    headers = ["CoA ID", "Raw Label", "Confidence", "Attribution / Rationale"]
    _write_header(ws, headers)
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["D"].width = 100
    learned = [m for m in coa_mappings if m.get("mapping_source") == "learned"]
    for r, m in enumerate(learned, 2):
        ws.cell(row=r, column=1, value=m.get("coa_id", "")).font = BODY_FONT
        ws.cell(row=r, column=2, value=m.get("raw_label", "")).font = BODY_FONT
        ws.cell(row=r, column=3, value=m.get("confidence")).font = BODY_FONT
        ws.cell(row=r, column=4, value=m.get("rationale", "")).font = BODY_FONT
    if not learned:
        ws.cell(row=2, column=1, value="(No learned mappings applied to this document.)").font = BODY_FONT


_RECON_STMT_LABEL = {
    "balance_sheet": "Balance Sheet",
    "income_statement": "Income Statement",
    "equity_statement": "Statement of Changes in Equity",
}
_RESULT = {True: "PASS", False: "FAIL", None: "INCOMPLETE"}


def _write_reconciliation_sheet(ws, reconciliation: dict) -> None:
    """Each extracted subtotal as a bold SUBTOTAL row (extracted vs computed,
    PASS/FAIL/INCOMPLETE, grouping method) followed by its component rows (mapping
    status, CoA, value, sign-flip). The 'Missing Leaves' flag counts unmapped
    components that are NOT themselves subtotals — the genuinely missing lines."""
    headers = ["Statement", "Kind", "Line Item", "Type / Status", "CoA ID",
               "CoA Name", "Confidence", "Value (latest yr)", "Computed Sum",
               "Difference", "Result", "Grouping", "Flags"]
    _write_header(ws, headers)
    for col, w in zip("ABCDEFGHIJKLM", [26, 13, 40, 16, 9, 24, 11, 15, 14, 14, 12, 11, 22]):
        ws.column_dimensions[col].width = w

    subtotals = (reconciliation or {}).get("subtotals", [])
    if not subtotals:
        ws.cell(row=2, column=1, value="(No subtotals extracted for this document.)").font = BODY_FONT
        return

    r = 2
    for sub in subtotals:
        yr = sub.get("headline_year")
        result = _RESULT[sub.get("pass")]
        comps = sub.get("components", [])
        missing = sum(1 for c in comps if c.get("status") != "mapped" and not c.get("is_subtotal"))
        flags = []
        if missing:
            flags.append(f"{missing} missing leaf line(s)")
        if any(c.get("sign_flipped") for c in comps):
            flags.append("sign flip")
        row_vals = [
            _RECON_STMT_LABEL.get(sub.get("statement_type"), sub.get("statement_type", "")),
            "▣ SUBTOTAL", sub.get("raw_label", ""), "Subtotal/Total", "", "", "",
            (sub.get("extracted") or {}).get(yr), (sub.get("computed") or {}).get(yr),
            (sub.get("difference") or {}).get(yr), result,
            sub.get("grouping_method", ""), "; ".join(flags),
        ]
        for c, v in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = SUBTOTAL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
        ws.cell(row=r, column=11).fill = {"PASS": _GREEN, "FAIL": _UNMAPPED_FILL,
                                          "INCOMPLETE": _AMBER}[result]
        r += 1

        for comp in comps:
            cflags = "SIGN FLIP" if comp.get("sign_flipped") else ""
            coa = f"{comp.get('coa_id', '')}".strip()
            row_vals = [
                "", "    └ component", comp.get("raw_label", ""), comp.get("status", ""),
                coa, comp.get("coa_name", ""), comp.get("confidence"),
                (comp.get("raw_values") or {}).get(yr), "", "", "", "", cflags,
            ]
            for c, v in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
            if comp.get("status") != "mapped" and not comp.get("is_subtotal"):
                ws.cell(row=r, column=4).fill = _UNMAPPED_FILL
            if comp.get("sign_flipped"):
                ws.cell(row=r, column=13).fill = _AMBER
            r += 1

    ws.freeze_panes = "A2"


def _write_usage_sheet(ws, usage: dict | None) -> None:
    """LLM token usage + estimated cost for the run, split by stage."""
    headers = ["Stage", "Calls", "Input Tokens", "Output Tokens",
               "Cache Read", "Cache Creation", "Est. Cost (USD)"]
    _write_header(ws, headers)
    for col, w in zip("ABCDEFG", [14, 8, 14, 14, 12, 14, 16]):
        ws.column_dimensions[col].width = w

    if not usage or not (usage.get("total") or {}).get("calls"):
        ws.cell(row=2, column=1,
                value="(No LLM usage recorded — offline export or no live run.)").font = BODY_FONT
        return

    r = 2
    for stage, b in (usage.get("by_stage") or {}).items():
        vals = [stage, b["calls"], b["input_tokens"], b["output_tokens"],
                b["cache_read"], b["cache_creation"], round(b["cost_usd"], 4)]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v).font = BODY_FONT
        r += 1
    t = usage["total"]
    tot = ["TOTAL", t["calls"], t["input_tokens"], t["output_tokens"],
           t["cache_read"], t["cache_creation"], round(t["cost_usd"], 4)]
    for c, v in enumerate(tot, 1):
        ws.cell(row=r, column=c, value=v).font = SUBTOTAL_FONT
    r += 2
    ws.cell(row=r, column=1, value=f"Models: {', '.join(usage.get('models') or [])}").font = BODY_FONT
    ws.cell(row=r + 1, column=1,
            value=usage.get("pricing_note", "Estimated at list price.")).font = BODY_FONT
    if usage.get("unknown_model_pricing"):
        ws.cell(row=r + 2, column=1,
                value="Note: some calls used a model with no pricing entry — cost understated.").font = BODY_FONT


def build_spread_xlsx(formatted: dict, coa_mappings: list[dict],
                      unmapped_items: list[dict] | None = None,
                      reconciliation: dict | None = None,
                      usage: dict | None = None) -> bytes:
    """Build the 6-sheet Spread workbook.

    Args:
        formatted:      output of spreading.spread_formatter.format_spread_output().
        coa_mappings:   the document's coa_mappings (for sheets 5 & 6).
        unmapped_items: the document's pending unmapped items (sheet 3). Defaults
                        to none — every extracted BS/P&L line is then either on the
                        statement sheets (mapped) or here (unmapped).
        reconciliation: the document's reconciliation_result (sheet 4). Defaults to
                        none — the sheet is still written with an empty-state note.
    """
    years = formatted.get("years", [])
    unmapped_items = unmapped_items or []

    # The unmapped sheet may carry year keys not present in the mapped set; widen
    # its columns to the union so no extracted value is silently hidden.
    unmapped_year_set = set(years)
    for item in unmapped_items:
        unmapped_year_set.update((item.get("value_spread") or {}).keys())
    unmapped_years = sorted(unmapped_year_set)

    wb = Workbook()

    ws_bs = wb.active
    ws_bs.title = "Balance Sheet"
    _write_statement_sheet(ws_bs, formatted.get("balance_sheet", []), years)

    ws_pl = wb.create_sheet("P&L")
    _write_statement_sheet(ws_pl, formatted.get("pl", []), years)

    ws_unmapped = wb.create_sheet("Unmapped Items")
    _write_unmapped_sheet(ws_unmapped, unmapped_items, unmapped_years)

    ws_recon = wb.create_sheet("Subtotal Reconciliation")
    _write_reconciliation_sheet(ws_recon, reconciliation or {})

    ws_conf = wb.create_sheet("Confidence & Source")
    _write_confidence_sheet(ws_conf, coa_mappings)

    ws_learn = wb.create_sheet("Learned Mappings Applied")
    _write_learned_sheet(ws_learn, coa_mappings)

    ws_usage = wb.create_sheet("Run Usage & Cost")
    _write_usage_sheet(ws_usage, usage)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
